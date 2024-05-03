import pytesseract
from PIL import Image, ImageEnhance
from openpyxl import Workbook
import os
import re
import cv2
import datetime

pytesseract.pytesseract.tesseract_cmd = r'c:/Users/Neel Amarnath Mulik/AppData/Local/Programs/Tesseract-OCR/tesseract.exe'

def extract_text_from_image(image_path):
    img = Image.open(image_path)
    
    # Enhance image for better OCR
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(1.5)
    
    img = img.resize((img.width * 2, img.height * 2), Image.LANCZOS)

    # Perform OCR
    extracted_text = pytesseract.image_to_string(img)
    
    confidence = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
    avg_confidence = sum(map(float, confidence['conf'])) / len(confidence['conf']) if confidence['conf'] else 0
    return extracted_text, avg_confidence

def find_name_licence_validity(text):
    name = ''
    licence_number = ''
    validity = ''
    
    # Define patterns for matching
    licence_keywords = ['licence number', 'dl no', 'dl number']
    name_keywords = ['name', 'driver name']
    validity_keywords = ['valid', 'expiry', 'valid till']

    for line in text.split('\n'):
        # Extract licence number
        if any(keyword in line.lower() for keyword in licence_keywords):
            match = re.search(r'([A-Za-z]{2}\s?-?\s?\d{2}\s?-?\s?\d{4}\s?-?\s?\d+)', line)  # Match license number pattern
            if match:
                licence_number = match.group().strip()

        # Extract name
        if any(keyword in line.lower() for keyword in name_keywords):
            match = re.search(r'(?:name|driver name)\s*[:=]?\s*(.*)', line, re.IGNORECASE)
            if match:
                name = match.group(1).strip()

        # Extract validity
        if any(keyword in line.lower() for keyword in validity_keywords):
            match = re.search(r':\s*(\d{2}-\d{2}-\d{4})', line)
            if match:
                validity = match.group(1).strip()

    return name, licence_number, validity


def extract_license_number(image_path):
    img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    template_img = cv2.imread(r"C:/Users/Neel Amarnath Mulik/Downloads/license-template.jpeg", cv2.IMREAD_GRAYSCALE)

    # Template Matching
    result = cv2.matchTemplate(img, template_img, cv2.TM_CCOEFF_NORMED)
    min_val, max_val, min_loc, max_loc = cv2.minMaxLoc(result)

    x, y = max_loc
    w, h = template_img.shape[::-1]
    license_region = img[y:y + h, x:x + w]

    # OCR on the extracted region
    extracted_text = pytesseract.image_to_string(license_region)
    return extracted_text.strip()

folder_path = r'C:/Users/Neel Amarnath Mulik/Mini4/mini-project-b/uploads'
image_files = [f for f in os.listdir(folder_path) if f.endswith(('.jpg', '.png', '.jpeg'))]

# Sort the image files numerically
image_files.sort(key=lambda x: int(re.search(r'\d+', x).group()))

# Initialize row index for writing to Excel
row_idx = 2
wb = Workbook()
ws = wb.active
ws.cell(row=1, column=1, value='Name')
ws.cell(row=1, column=2, value='Licence Number')
ws.cell(row=1, column=3, value='Validity')
ws.cell(row=1, column=4, value='File name')  # Add a new column for file names

for image_file in image_files:
    image_path = os.path.join(folder_path, image_file)

    extracted_text, confidence = extract_text_from_image(image_path)

    if confidence < 40:
        print(f"Warning: Image {image_file} may be blurred or extraction failed (confidence: {confidence})")
        print(f"Upload another copy of the image")
        continue

    name, licence_number, validity = find_name_licence_validity(extracted_text)

    # Refine licence number extraction using template matching
    if not licence_number:
        licence_number = extract_license_number(image_path)

    # Convert validity to datetime object
    validity_date = datetime.datetime.strptime(validity, '%d-%m-%Y') if validity else None

    ws.cell(row=row_idx, column=1, value=name)
    ws.cell(row=row_idx, column=2, value=licence_number)
    ws.cell(row=row_idx, column=3, value=validity_date)
    ws.cell(row=row_idx, column=4, value=image_file)

    row_idx += 1

output_file = r'c:/Users/Neel Amarnath Mulik/Downloads/license_det.xlsx'
wb.save(output_file)

print(f"License information extracted and saved to {output_file}")
